from decimal import Decimal

import openpyxl
import pandas as pd
from numpy import datetime64
from openpyxl import load_workbook, Workbook
import Func

# EarliestTime = datetime64("2000-01-02")  # 设置工艺路线版本日期的最早期限
order = ['单据号码', '单据日期', '制单人', '报工单表头备注', '项目号', '生产订单', '行号', '物料编码', '物料名称', '工序行号', '标准工序',
         '工序名称', '生产数量', '合格数量', '班组编码', '班组名称', '员工代号', '员工姓名', '工作中心', '工作中心名称', '备注设备',
         '备注工时', '版本代号', '单位标准工时', '总标准工时', '实际报工工时']


# 对比 当月的 实际报工工时，大于0 返回 资源工时1 ，否则 返回 资源工时2
def FirsSecondChoice(df):
    if df['实际报工工时'] > 0:
        return df['资源工时1']
    else:
        return df['资源工时2']


# # 计划工时分组计算合并值
# def GetSumPlanData(Data):
#     sum_data_list = []
#     for name, group in Data.groupby(["物料编码"]):
#         qualified_num = group['工时(分子)'].sum()  # 取合格数量总值保留两位
#         # qualified_num = Decimal(qualified_num).quantize(Decimal('0.00'))
#         group.loc[:, "计划工时"] = qualified_num  # 新建 总合格数量 列
#         sum_data_list.append(group.head(1))
#     return sum_data_list


# 获取最大版本代号
def GetMaxVersionNumber(Data):
    max_data_list = []
    for name, group in Data.groupby(["物料编码"]):
        group = pd.DataFrame(group)  # 新建pandas
        group = group.sort_values(by='版本代号', ascending=False)  # 降序排序
        MaxNumber = group.iloc[0, 1]  # 取降序排序第一个 版本代号
        group = group[group["版本代号"] == MaxNumber]  # 筛选 版本代号 最大的数据
        max_data_list.append(group)
    return pd.concat(max_data_list)


# # 实际工时分组计算合并值
# def GetSumActualData(Data):
#     sum_data_list = []
#     for name, group in Data.groupby(["物料编码", "生产订单", "行号"]):
#         qualified_num = group['实际报工工时'].sum()  # 取合格数量总值保留两位
#         group.loc[:, "实际报工工时"] = qualified_num  # 新建 总合格数量 列
#         sum_data_list.append(group.head(1))
#     return sum_data_list


class WorkHour:
    def __init__(self):
        self.WeldingPart = []  # 铆焊车间List
        self.Assembly = []  # 总装车间List
        self.AssemblyPX = []  # PX车间List
        self.AssemblyECC = []  # 电控柜车间List
        self.Machining = []  # 机加工车间List

        self.WeldingPart_FX = []  # 铆焊车间返修List
        self.Assembly_FX = []  # 总装车间返修List
        self.AssemblyPX_FX = []  # PX车间返修List
        self.AssemblyECC_FX = []  # 电控柜车间返修List
        self.Machining_FX = []  # 机加工车间返修List

        self.wb = Workbook()

        self.func = Func
        self.ThisMonthStart, self.ThisMonthEnd, self.LastMonthEnd, self.LastMonthStart = self.func.GetDate()
        self.path = Func.Path()
        self.work_list = ['打磨', '大、小件划线', '焊接', '冷作', '修毛刺', '人工工时', '水压试验', '攻螺纹、修毛刺、清洁', '清洁']

        self.PlanHourData = pd.DataFrame
        self.ProductionData = pd.read_excel(f"{self.path}/DATA/PROD/生产订单列表.XLSX",
                                            usecols=['生产订单号', '行号', '物料编码', '工艺路线版本号'],  # '物料名称', '生产批号', '制单时间', '类型',
                                            converters={'生产订单号': str, '行号': str, '物料编码': str, '工艺路线版本号': int})  # '制单时间': # datetime64,  '生产批号': str,
        self.ProductionData = self.ProductionData.rename(columns={'生产订单号': '生产订单', '工艺路线版本号': '版本代号'})  #

        self.Routing_data = pd.read_excel(f"{self.path}/DATA/SCM/OM/工艺路线资料表--含资源.xlsx",
                                          header=3, usecols=["物料编码", "工作中心", "工作中心名称", "版本日期", "版本代号", "资源名称",
                                                             "版本说明", "工时(分子)", "工序行号", "工序代号"],
                                          converters={'物料编码': str, "工时(分子)": int, "版本代号": int, "工序行号": str})

        self.WorkHourData = pd.read_excel(f"{self.path}/DATA/PROD/报工列表.xlsx",
                                          usecols=["单据日期", "单据号码", "制单人", "表头备注", "生产批号", "生产订单", "行号",
                                                   "物料编码", "物料名称", "生产数量", "资源工时1", "资源名称1",
                                                   "资源工时2", "资源名称2", "移入工序行号", "移入标准工序", "移入工序说明", "合格数量",
                                                   "班组编码", "班组名称", "员工代号", "员工姓名", "备注设备", "备注工时"],
                                          converters={'行号': str, "资源工时1": int, "资源工时2": int, "备注工时": int,
                                                      "生产数量": int, '物料编码': str, '移入工序行号': str,
                                                      '员工代号': str, "表头备注": str})
        self.WorkHourData = self.WorkHourData.rename(columns={'移入工序行号': '工序行号', '移入标准工序': '标准工序', '移入工序说明': '工序名称',
                                                              '生产批号': '项目号', '表头备注': '报工单表头备注'})
        self.WorkHourData = pd.merge(self.WorkHourData, self.ProductionData, how="left", on=['生产订单', '行号', '物料编码'])

        self.Routing_data = self.Routing_data.rename(columns={'工序代号': '标准工序', '工时(分子)': '单位标准工时'})

        # 重新定义 版本日期格式，再转化为 datatime64
        self.Routing_data = self.Routing_data.dropna(subset=['物料编码'])  # 去除nan的列
        self.Routing_data["版本日期"] = self.Routing_data["版本日期"].str.replace("/", "-").astype("datetime64")

        # 筛选 资源名称1，资源名称2 的符合work_list 的 资源名称
        WorkHourData_First = self.WorkHourData[self.WorkHourData['资源名称1'].isin(self.work_list)]
        WorkHourData_Second = self.WorkHourData[self.WorkHourData['资源名称2'].isin(self.work_list)]
        self.Routing_data = self.Routing_data[self.Routing_data['资源名称'].isin(self.work_list)]
        self.WorkHourData = pd.merge(WorkHourData_First, WorkHourData_Second, how="outer")
        self.WorkHourData["实际报工工时"] = self.WorkHourData['资源工时1'] - self.WorkHourData['资源工时2']
        # 放入函数 FirsSecondChoice 对比取值
        self.WorkHourData["实际报工工时"] = self.WorkHourData.apply(FirsSecondChoice, axis=1)

        self.RepairData = self.Routing_data[self.Routing_data["版本说明"] == "返修"]
        self.StandardData = self.Routing_data[self.Routing_data["版本说明"] != "返修"]
        # self.RepairData = GetMaxVersionNumber(self.RepairData)  # 获取版本最大值
        # self.StandardData = GetMaxVersionNumber(self.StandardData)  # 获取版本最大值

        self.RepairProductData = self.WorkHourData.loc[self.WorkHourData['生产订单'].str.contains(r'^F')]  # 模糊查询以F开头的
        self.ProductData = self.WorkHourData.loc[self.WorkHourData['生产订单'].str.contains(r'^(?!F).*')]  # 模糊查询不以F开头的

    def GetWorkData(self, group):

        group["总标准工时"] = group["合格数量"] * group["单位标准工时"]  # 计算 标准工时
        del group["资源名称"]
        # del group["版本代号"]
        del group["版本日期"]
        del group["资源工时1"]
        del group["资源工时2"]
        del group["资源名称1"]
        del group["资源名称2"]
        # group = group.rename(columns={'生产批号': '项目号'})
        group = group[order]
        group = group.sort_values(by=['生产订单', '行号', '物料编码', '工序行号', '标准工序'], ascending=True)  # 升序排列
        return group

    def GetWorkHour(self):
        # RepairData.to_excel(f'{self.path}/RESULT/WORKHOUR/RepairData.xlsx', sheet_name="0", index=False)
        # # 使用正则表达式尽心模糊匹配
        # PXData = self.Routing_data.loc[self.Routing_data['工作中心'].str.contains('^P')]  # 模糊查询总装PX 车间

        for name, group in self.ProductData.groupby(["制单人"]):  # 按车间操作人员进行拆分  正常生产订单部分
            group = pd.DataFrame(group)  # 新建pandas
            if name == "郭东升":
                # 对各个车间的 以物料编码 为 主键 将 工时(分子) 进行sum合计，返回的值进行合并
                group = pd.merge(group, self.StandardData, on=['物料编码', '工序行号', '标准工序', '版本代号'])  #
                group = self.GetWorkData(group)
                self.AssemblyPX.append(group)
            elif name == "黄鑫凯":
                # Data = pd.concat(GetSumPlanData(self.Routing_data), axis=0, ignore_index=True)
                group = pd.merge(group, self.StandardData, on=['物料编码', '工序行号', '标准工序', '版本代号'])
                group = self.GetWorkData(group)
                self.AssemblyECC.append(group)
            elif name == "乐美珠":
                # Data = pd.concat(GetSumPlanData(self.Routing_data), axis=0, ignore_index=True)
                group = pd.merge(group, self.StandardData, on=['物料编码', '工序行号', '标准工序', '版本代号'])
                group = self.GetWorkData(group)
                self.Machining.append(group)
            elif name == "吕春华" or name == "夏正棋":
                # Data = pd.concat(GetSumPlanData(self.Routing_data), axis=0, ignore_index=True)
                group = pd.merge(group, self.StandardData, on=['物料编码', '工序行号', '标准工序', '版本代号'])
                group = self.GetWorkData(group)
                self.Assembly.append(group)
            elif name == "杨薇1" or name == "林李旭":
                # Data = pd.concat(GetSumPlanData(self.Routing_data), axis=0, ignore_index=True)
                group = pd.merge(group, self.StandardData, on=['物料编码', '工序行号', '标准工序', '版本代号'])
                group = self.GetWorkData(group)
                self.WeldingPart.append(group)

        for name, group in self.RepairProductData.groupby(["制单人"]):  # 按车间操作人员进行拆分 返修部分
            group = pd.DataFrame(group)  # 新建pandas
            if name == "郭东升":
                # 对各个车间的 以物料编码 为 主键 将 工时(分子) 进行sum合计，返回的值进行合并
                group = pd.merge(group, self.RepairData, on=['物料编码', '工序行号', '标准工序', '版本代号'])
                group = self.GetWorkData(group)
                self.AssemblyPX_FX.append(group)
            elif name == "黄鑫凯":
                group = pd.merge(group, self.RepairData, on=['物料编码', '工序行号', '标准工序', '版本代号'])
                group = self.GetWorkData(group)
                self.AssemblyECC_FX.append(group)
            elif name == "乐美珠":
                group = pd.merge(group, self.RepairData, on=['物料编码', '工序行号', '标准工序', '版本代号'])
                group = self.GetWorkData(group)
                self.Machining_FX.append(group)
            elif name == "吕春华" or name == "夏正棋":
                group = pd.merge(group, self.RepairData, on=['物料编码', '工序行号', '标准工序', '版本代号'])
                group = self.GetWorkData(group)
                self.Assembly_FX.append(group)
            elif name == "杨薇1" or name == "林李旭":
                group = pd.merge(group, self.RepairData, on=['物料编码', '工序行号', '标准工序', '版本代号'])
                group = self.GetWorkData(group)
                self.WeldingPart_FX.append(group)

    def mkdir(self, path):
        self.func.mkdir(path)

    def SaveSheet(self, data, name):  # 新建excel页签并保存数据
        SaveBook = load_workbook(f'{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx')
        writer = pd.ExcelWriter(f"{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx", engine='openpyxl')
        writer.book = SaveBook
        data.to_excel(writer, f"{name}", index=False)
        writer.save()

    def FullSaveSheet(self, name):  # 新建空excel页签并保存数据
        data = pd.DataFrame()
        SaveBook = load_workbook(f'{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx')
        writer = pd.ExcelWriter(f"{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx", engine='openpyxl')
        writer.book = SaveBook
        data.to_excel(writer, f"{name}", index=False)
        writer.save()

    def SaveFile(self):
        # self.AsNameList[0] = self.AsNameList[0][order]

        self.wb.save(f'{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx')
        try:
            self.SaveSheet(pd.concat(self.AssemblyECC, axis=0, ignore_index=True), "总装电控")
        except:
            self.FullSaveSheet("总装电控")
        try:
            self.SaveSheet(pd.concat(self.AssemblyPX, axis=0, ignore_index=True), "总装PX")
        except:
            self.FullSaveSheet("总装PX")
        try:
            self.SaveSheet(pd.concat(self.Machining, axis=0, ignore_index=True), "机加工")
        except:
            self.FullSaveSheet("机加工")
        try:
            self.SaveSheet(pd.concat(self.Assembly, axis=0, ignore_index=True), "总装")
        except:
            self.FullSaveSheet("总装")
        try:
            self.SaveSheet(pd.concat(self.WeldingPart, axis=0, ignore_index=True), "铆焊")
        except:
            self.FullSaveSheet("铆焊")
        try:
            self.SaveSheet(pd.concat(self.AssemblyECC_FX, axis=0, ignore_index=True), "总装电控返修")
        except:
            self.FullSaveSheet("总装电控返修")
        try:
            self.SaveSheet(pd.concat(self.Machining_FX, axis=0, ignore_index=True), "机加工返修")
        except:
            self.FullSaveSheet("机加工返修")
        try:
            self.SaveSheet(pd.concat(self.Assembly_FX, axis=0, ignore_index=True), "总装返修")
        except:
            self.FullSaveSheet("总装返修")
        try:
            self.SaveSheet(pd.concat(self.WeldingPart_FX, axis=0, ignore_index=True), "铆焊返修")
        except:
            self.FullSaveSheet("铆焊返修")
        try:
            self.SaveSheet(pd.concat(self.AssemblyPX_FX, axis=0, ignore_index=True), "总装PX返修")
        except:
            self.FullSaveSheet("总装PX返修")

        workbook = openpyxl.load_workbook(f'{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx')
        del workbook["Sheet"]
        workbook.save(f'{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx')

    def run(self):
        self.GetWorkHour()
        self.func.mkdir(self.path + '/RESULT/WORKHOUR')
        self.SaveFile()
        # del self.wb["Sheet"]
        # self.wb.save(f'{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx')


if __name__ == '__main__':
    WH = WorkHour()
    WH.run()

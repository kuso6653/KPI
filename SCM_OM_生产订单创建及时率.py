import pandas as pd
from numpy import datetime64
from openpyxl import load_workbook, Workbook
import re
import Func

EarliestTime = datetime64("2000-01-02")  # 设置工艺路线版本日期的最早期限

from OracleHelper import OracleHelper


# -*- coding:utf-8 -*-
def HaveLog(strRE):  # 判断该行是否有等号
    EqualRe = re.findall("(^=)", strRE)
    try:
        if EqualRe[0] == "=":
            return True
    except:
        return False


def HaveBOM(strRE):  # 判断该行是否有BOM字段
    BOMRe = re.findall("[F][u][n][T][y][p][e][:][B][O][M]", strRE)
    try:
        if BOMRe[0] == "FunType:BOM":
            return True
    except:
        return False


def HaveRow(strRE):  # 判断该行是否有iRowNo字段
    RowRe = re.findall("[i][R][o][w][N][o]", strRE)
    try:
        if RowRe[0] == "iRowNo":
            return True
    except:
        return False


class OrderCreation:
    def __init__(self):
        self.func = Func
        self.ThisMonthStart, self.ThisMonthEnd, self.LastMonthEnd, self.LastMonthStart = self.func.GetDate()
        self.path = Func.Path()
        # 将上月首尾日期切割
        self.OtherMonthEnd = str(self.ThisMonthEnd).split(" ")[0].replace("", "")  # 取当月最后一天
        self.ThisMonthStart = str(self.ThisMonthStart).split(" ")[0].replace("-", "")
        self.ThisMonthEnd = str(self.ThisMonthEnd).split(" ")[0].replace("-", "")

        self.ProductionData = pd.read_excel(
            f"{self.path}/DATA/PROD/生产订单列表.XLSX",
            usecols=['生产订单号', '行号', '物料编码', '物料名称', '生产批号', '制单时间', '类型'],
            converters={'生产订单号': str, '物料编码': str, '生产批号': str})
        self.ProductionData['制单时间'] = self.ProductionData['制单时间'].astype(datetime64)
        self.ProductionData = self.ProductionData.dropna(subset=['生产批号'])  # 去除nan的列
        self.ProductionData['生产批号'] = self.ProductionData['生产批号'].str[:5]  # 截取前五位
        self.Material_data = pd.read_excel(f"{self.path}/DATA/SCM/存货档案{self.ThisMonthEnd}.XLSX",
                                           usecols=['存货编码', '计划默认属性', '启用日期'],
                                           converters={'启用日期': datetime64, "存货编码": str})

        self.Routing_data = pd.read_excel(f"{self.path}/DATA/SCM/OM/工艺路线资料表--含资源.xlsx",
                                          usecols=[0, 4, 6], header=3, names=["物料编码", "版本代号", "版本日期"],
                                          converters={'版本日期': str, '物料编码': str})

        # self.BOM_data = pd.read_excel(f"{self.path}/DATA/SCM/OM/BOM集成时间表.xlsx", header=1,
        #                               usecols=['子件编码', '计划默认属性', '集成时间'],
        #                               converters={"子件编码": str})
        bom = BOM()
        self.BOM_data = bom.run()


    def mkdir(self, path):
        self.func.mkdir(path)

    def ModifyFormat(self):  # 修改数据基本格式
        self.ProductionData = self.ProductionData[self.ProductionData["类型"] == "标准"]
        self.Material_data = self.Material_data.rename(columns={'存货编码': '物料编码'})
        self.Routing_data = self.Routing_data.dropna(subset=['版本代号'])  # 去除nan的列
        self.Routing_data["版本日期"] = self.Routing_data["版本日期"].str.replace("/", "-").astype("datetime64")
        self.Routing_data = self.Routing_data[self.Routing_data["版本日期"] > EarliestTime]
        self.Routing_data = self.Routing_data.drop_duplicates(subset=["物料编码"])  # 去重
        self.ProductionData = self.ProductionData[self.ProductionData["制单时间"] > self.ThisMonthStart]
        self.Material_data = self.Material_data[self.Material_data["启用日期"] > self.ThisMonthStart]
        # 合并生成当月数据在导入到U8中进行再查询
        self.ThisMonthData = pd.merge(self.ProductionData, self.Material_data, how="left", on=['物料编码'])
        # self.ThisMonthData.to_excel(f'{self.path}/RESULT/SCM/OM/当月物料查询表.xlsx', sheet_name="当月物料查询表", index=False)
        # 当月 启用日期 为空的物料无旧物料
        self.OldMaterialData = self.ThisMonthData[self.ThisMonthData["启用日期"].isnull()]  # 旧物料
        del self.OldMaterialData['计划默认属性']
        del self.OldMaterialData['启用日期']
        del self.BOM_data["cProNo"]
        self.BOM_data = self.BOM_data.rename(
            columns={'cFaInvCode': '物料编码', 'time': '集成时间', 'name': '生产批号', 'customername': '客户名称'})
        self.BOM_data["集成时间"] = self.BOM_data["集成时间"].astype("datetime64")

    def GetNewMaterial(self):  # 获取新物料
        self.mkdir(self.path + "/RESULT/SCM/OM")
        NewMaterialData = self.ThisMonthData[self.ThisMonthData["启用日期"].notnull()]  # 新物料
        NewProductionData = pd.merge(NewMaterialData, self.Routing_data, on=["物料编码"], how="left")
        NewProductionData = NewProductionData.dropna(subset=["版本日期"])  # 去nan
        NewProductionData['下单延时/H'] = (
                    (NewProductionData['制单时间'] - NewProductionData['版本日期']) / pd.Timedelta(1, 'H')).astype(int)
        try:
            NewProductionData.loc[NewProductionData["下单延时/H"] > 72, "创建及时率"] = "超时"  # 计算出来的审批延时大于3天为超时
            NewProductionData.loc[NewProductionData["下单延时/H"] <= 72, "创建及时率"] = "正常"  # 小于等于3天为正常

            try:
                NewProductionCount = NewProductionData['创建及时率'].value_counts()['超时']
            except:
                NewProductionCount = 0

            NewProductionCountAll = NewProductionData.shape[0]
            NewProductionResult = format(float(1 - NewProductionCount / NewProductionCountAll), '.2%')
            dict = {'当月未及时创建生产订单新物料数': [NewProductionCount], '当月已创建生产订单新物料总数': [NewProductionCountAll],
                    '生产订单新物料创建及时率': [NewProductionResult]}
            NewProductionResult_sheet = pd.DataFrame(dict)

            # 输出新物料及时率
            NewProductionResult_sheet.to_excel(f'{self.path}/RESULT/SCM/OM/生产订单创建及时率.xlsx', sheet_name="生产订单新物料创建及时率", index=False)
            book = load_workbook(f'{self.path}/RESULT/SCM/OM/生产订单创建及时率.xlsx')
            writer = pd.ExcelWriter(f"{self.path}/RESULT/SCM/OM/生产订单创建及时率.xlsx", engine='openpyxl')
            writer.book = book
            NewProductionData.to_excel(writer, "生产订单新物料创建情况", index=False)
            writer.save()

        except:
            df = pd.DataFrame()
            NewProductionCount = 0
            NewProductionCountAll = 0
            NewProductionResult = 0
            dict = {'当月未及时创建生产订单新物料数': [NewProductionCount], '当月已创建生产订单新物料总数': [NewProductionCountAll],
                    '生产订单新物料创建及时率': [NewProductionResult]}
            NewProductionResult_sheet = pd.DataFrame(dict)
            NewProductionResult_sheet.to_excel(f'{self.path}/RESULT/SCM/OM/生产订单创建及时率.xlsx', sheet_name="生产订单新物料创建及时率", index=False)
            book = load_workbook(f'{self.path}/RESULT/SCM/OM/生产订单创建及时率.xlsx')
            writer = pd.ExcelWriter(f"{self.path}/RESULT/SCM/OM/生产订单创建及时率.xlsx", engine='openpyxl')
            writer.book = book
            df.to_excel(writer, "生产订单新物料创建情况", index=False)
            writer.save()

    def GetOldMaterial(self):  # 获取旧物料

        OldProductionData = pd.merge(self.OldMaterialData, self.BOM_data, on=["物料编码", "生产批号"], how="left")
        OldProductionData = OldProductionData.dropna(subset=["集成时间"])  # 去nan
        OldProductionData['下单延时/H'] = (
                (OldProductionData['制单时间'] - OldProductionData['集成时间']) / pd.Timedelta(1, 'H')).astype(
            int)
        OldProductionData = OldProductionData.drop_duplicates(subset=["生产订单号", "行号", "物料编码"])  # 去重
        OldProductionData.loc[OldProductionData["下单延时/H"] > 24, "创建及时率"] = "超时"  # 计算出来的审批延时大于1天为超时
        OldProductionData.loc[OldProductionData["下单延时/H"] <= 24, "创建及时率"] = "正常"  # 小于等于1天为正常

        try:
            OldProductionCount = OldProductionData['创建及时率'].value_counts()['超时']
        except:
            OldProductionCount = 0

        OldProductionCountAll = OldProductionData.shape[0]
        OldProductionResult = format(float(1 - OldProductionCount / OldProductionCountAll), '.2%')
        dict = {'当月未及时创建生产订单旧物料数': [OldProductionCount], '当月已创建生产订单旧物料总数': [OldProductionCountAll],
                '生产订单旧物料创建及时率': [OldProductionResult]}
        OldProductionResult_sheet = pd.DataFrame(dict)

        # 输出旧物料及时率
        self.mkdir(self.path + "/RESULT/SCM/OM")
        book = load_workbook(f'{self.path}/RESULT/SCM/OM/生产订单创建及时率.xlsx')
        writer = pd.ExcelWriter(f"{self.path}/RESULT/SCM/OM/生产订单创建及时率.xlsx", engine='openpyxl')
        writer.book = book
        OldProductionResult_sheet.to_excel(writer, "生产订单旧物料创建及时率", index=False)
        OldProductionData.to_excel(writer, "生产订单旧物料创建情况", index=False)
        writer.save()

    def run(self):
        self.ModifyFormat()
        self.GetNewMaterial()
        self.GetOldMaterial()

# 分析当月BOMlog的日期抓取
class BOM:
    def __init__(self):
        self.func = Func
        self.ThisMonthStart, self.ThisMonthEnd, self.LastMonthEnd, self.LastMonthStart = self.func.GetDate()
        self.path = Func.Path()
        self.BOMList = []
        self.oneStr = 1
        self.cursor = 0  # 设置游标

        self.BOMData = pd.DataFrame(columns=['cProNo', 'cFaInvCode', 'time'])
        self.TagTime = ""
        # self.wb = Workbook()
        # self.wb.save('./BOM.xlsx')
        sql_PRO = "select t.code,t.name,t.customername from TN_A_PROJECT t"
        SqlOracle = OracleHelper("T5_ENTITY", "thsoft", "10.56.164.22:1521/THPLM")
        sql_PRO_list = SqlOracle.find_sql(sql_PRO)  # 抓取数据库函数
        self.PROData = pd.DataFrame(sql_PRO_list, columns=['cProNo', 'name', 'customername'])

    def getTXT(self, lines, this_date):
        flag = 1
        for index, value in enumerate(lines):
            if HaveLog(value) and flag == 0:  # 有等号并且第一次遇到
                self.cursor = 0
                #  = pd.DataFrame(columns=['cProNo', 'cFaInvCode', 'time'])
                flag = 1  # 设置第二次遇到直接跳过直到遇到BOM 字段
            if HaveBOM(value):  # 有BOM
                self.cursor = 1  # 将游标置为 1
            if self.cursor == 1:
                self.TagTime = lines[index - 2]
                self.cursor = 2
            if self.cursor == 2:
                flag = 0
                if HaveRow(value):
                    self.BOMData = self.BOMData.append({"cProNo": lines[index + 1].replace('cProNo', '').replace(',',
                                                                                                                 '').replace(
                        '"', '').replace(':', '').replace(' ', '').replace('\n', ''),
                                                        "cFaInvCode": lines[index + 2].replace('cFaInvCode',
                                                                                               '').replace(',',
                                                                                                           '').replace(
                                                            '"', '').replace(':', '').replace(' ', '').replace('\n',
                                                                                                               ''),
                                                        "time": self.TagTime.replace('时间:', '').replace('\n', '')},
                                                       ignore_index=True)

    def run(self):
        self.ThisMonthStart = str(self.ThisMonthStart).split(" ")[0]
        this_month = self.ThisMonthStart.split("-")[1]
        year = self.ThisMonthStart.split("-")[0]
        month = self.ThisMonthStart.split("-")[1]
        EveryDays = self.func.WorkDays(year, this_month)  # 取本月所有日期
        EveryDays = self.func.ReformDays(EveryDays)  # 改造

        for i in EveryDays:
            this_date = str(year) + str(month) + str(i)
            try:
                FileOpen = open(f'{self.path}/DATA/SCM/OM/U8接口{this_date}_u8log.txt', "rt", encoding="utf-8")
                lines = FileOpen.readlines()
                self.getTXT(lines, this_date)
            except:
                continue
        self.BOMData = pd.merge(self.BOMData, self.PROData, how="left", on="cProNo")  # 左链接
        max_data_list = []
        for name, group in self.BOMData.groupby(["cProNo", "cFaInvCode"]):
            group = pd.DataFrame(group)  # 新建pandas
            group = group.sort_values(by='time', ascending=False)  # 降序排序
            max_data_list.append(group.head(1))
        self.BOMData = pd.concat(max_data_list, axis=0, ignore_index=True)
        # self.BOMData.to_excel('./BOM.xlsx', index=False)
        return self.BOMData


if __name__ == '__main__':
    OC = OrderCreation()
    OC.run()

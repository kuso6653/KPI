import datetime
from datetime import time

import pandas as pd
from numpy import datetime64
from openpyxl import load_workbook

import Func


class Warehouse:
    def __init__(self):
        WarehouseList = ['机电库', '钢材库', '油料库', '辅料库', '标准件库', '工具劳保库', '外购件库(毛坯）',
                         '外协件库', 'KM电控柜库', '塑机库', '电控柜客供料仓库', '型材外协库', '钢板外协库', 'PX及电控柜半成品库',
                         ]
        self.func = Func
        self.ThisMonthStart, self.ThisMonthEnd, self.LastMonthEnd, self.LastMonthStart = self.func.GetDate()
        self.path = Func.Path()

        # 将本月月首尾日期切割
        self.LastMonthStart = str(self.LastMonthStart).split(" ")[0].replace("-", "")
        self.ThisMonthStart = str(self.ThisMonthStart).split(" ")[0].replace("-", "")
        self.ThisMonthEnd = str(self.ThisMonthEnd).split(" ")[0].replace("-", "")
        self.PurchaseInData = pd.read_excel(
            f"{self.path}/DATA/SCM/采购时效性统计表.XLSX",
            usecols=[1, 6, 7, 12, 23, 27, 29, 30, 31], header=2,
            names=["订单号", "存货编码", "存货名称", "订单审核时间", "报检审核时间", "检验审核时间", '入库单号', '行号', "入库制单时间"],
            converters={'订单审核时间': datetime64, '报检审核时间': datetime64,
                        '检验审核时间': datetime64, '入库制单时间': datetime64,
                        '存货编码': float, '入库单号': str})

        self.MaterialInData = pd.read_excel(
            f"{self.path}/DATA/SCM/WM/采购入库单列表.XLSX",
            usecols=['仓库', '入库单号', '行号'],
            converters={'入库单号': str})

        self.MaterialInData = self.MaterialInData[
            self.MaterialInData['仓库'].isin(WarehouseList)]

        self.MaterialOutData = pd.read_excel(
            f"{self.path}/DATA/SCM/WM/材料出库单列表.XLSX",
            usecols=['出库单号', '材料编码', '物料描述', '行号', '仓库'],
            converters={'材料编码': str, '出库单号': str})
        self.MaterialOutData = self.MaterialOutData[
            self.MaterialOutData['仓库'].isin(WarehouseList)]

        self.WorkFlowData = pd.read_excel(
            f"{self.path}/DATA/SCM/WM/工作流处理追溯.XLSX",
            usecols=['单据编号', '处理人', '处理时间', '处理动作'], header=1,
            converters={'单据编号': str, '处理时间': datetime64, '处理动作': str})
        self.WorkFlowList = []

    def mkdir(self, path):
        self.func.mkdir(path)

    def GetWarehouse(self):

        # 采购入库及时率
        self.MaterialInData = self.MaterialInData.dropna(axis=0, how='any')  # 去除所有nan的列
        self.PurchaseInData = self.PurchaseInData.dropna(axis=0, how='any')  # 去除所有nan的列
        # 采购入库单列表 和 采购时效性统计表 合并
        self.PurchaseInData = pd.merge(self.PurchaseInData, self.MaterialInData, on=['入库单号', '行号'])
        self.PurchaseInData['审批时间/H'] = ((self.PurchaseInData['入库制单时间'] - self.PurchaseInData['检验审核时间'])
                                         / pd.Timedelta(1, 'H')).astype(int)  # 制单时间相减，然后减去 质检的审核时间
        # 将天数转化为小时数
        self.PurchaseInData.loc[self.PurchaseInData["审批时间/H"] > 72, "单据状态"] = "超时"  # 计算出来的审批延时大于72为超时
        self.PurchaseInData.loc[self.PurchaseInData["审批时间/H"] <= 72, "单据状态"] = "正常"  # 小于等于72为正常

        try:
            PurchaseInCount = self.PurchaseInData['单据状态'].value_counts()['超时']
        except:
            PurchaseInCount = 0

        PurchaseInCountAll = len(self.PurchaseInData)
        PurchaseInResult = format(float(1 - PurchaseInCount / PurchaseInCountAll), '.2%')
        dict1 = {'当月未及时入库物料数': [PurchaseInCount], '当月入库物料总数': [PurchaseInCountAll],
                 '仓库入库及时率': [PurchaseInResult]}
        PurchaseInResult_sheet = pd.DataFrame(dict1)

        # 材料出库及时率
        # 将 WorkFlowData 的数据分组保存，
        # 取分组后最后降序排列
        # 取第一个也就是最大时间
        for name1, group in self.WorkFlowData.groupby(["单据编号"]):
            group = pd.DataFrame(group)  # 新建pandas
            group = group.sort_values(by='处理时间', ascending=False)  # 降序排序
            # 筛选 处理动作 不为 同意 或者 提交的
            if group.iloc[0]["处理动作"] in ["同意", "提交"]:
                group['审批时间/H'] = group['处理时间'] - group['处理时间'].shift(-1)
                group['上一流程处理时间'] = group['处理时间'].shift(-1)
                self.WorkFlowList.append(group.head(1))
        self.WorkFlowData = pd.concat(self.WorkFlowList, axis=0, ignore_index=True)
        self.WorkFlowData = self.WorkFlowData.rename(columns={'单据编号': '出库单号'})
        self.MaterialOutData = pd.merge(self.MaterialOutData, self.WorkFlowData, on=["出库单号"])
        self.MaterialOutData = self.MaterialOutData.dropna(axis=0, how='any')  # 去除所有nan的列
        self.MaterialOutData["审批时间/H"] = (
                (self.MaterialOutData["审批时间/H"]) / pd.Timedelta(1, 'H')).astype(int)  # 转化为int小时
        # self.MaterialOutData["审批时间/H"] = self.MaterialOutData["审批时间/H"].astype(int)
        self.MaterialOutData.reset_index()
        self.MaterialOutData.loc[self.MaterialOutData["审批时间/H"] > 72, "单据状态"] = "超时"  # 计算出来的审批延时大于72为超时
        self.MaterialOutData.loc[self.MaterialOutData["审批时间/H"] <= 72, "单据状态"] = "正常"  # 小于等于72为正常
        self.MaterialOutData = self.MaterialOutData.drop_duplicates()  # 去重
        try:
            MaterialOutCount = self.MaterialOutData['单据状态'].value_counts()['超时']
        except:
            MaterialOutCount = 0

        MaterialOutCountAll = len(self.MaterialOutData)
        MaterialOutResult = format(float(1 - MaterialOutCount / MaterialOutCountAll), '.2%')
        dict2 = {'当月未及时出库物料数': [MaterialOutCount], '当月出库物料总数': [MaterialOutCountAll],
                 '仓库出库及时率': [MaterialOutResult]}
        MaterialOutResult_sheet = pd.DataFrame(dict2)
        self.PrintOut(PurchaseInResult_sheet, MaterialOutResult_sheet)

    def PrintOut(self, PurchaseInResult_sheet, MaterialOutResult_sheet):  # 仓库出入库及时率输出
        del self.PurchaseInData['订单号']
        del self.PurchaseInData['订单审核时间']
        del self.PurchaseInData['报检审核时间']
        del self.MaterialOutData['处理动作']

        self.PurchaseInData = self.PurchaseInData.rename(columns={'行号': '入库单行号', '仓库': '入库仓库'})
        order = ['入库单号', '入库单行号', '入库仓库', '存货编码', '存货名称', '检验审核时间', '入库制单时间', '审批时间/H', '单据状态']
        self.PurchaseInData = self.PurchaseInData[order]

        self.MaterialOutData = self.MaterialOutData.rename(columns={'行号': '材料出库单行号', '仓库': '出库仓库'})
        order = ['出库单号', '材料出库单行号', '出库仓库', '材料编码', '物料描述', '处理人', '上一流程处理时间', '处理时间', '审批时间/H', '单据状态']
        self.MaterialOutData = self.MaterialOutData[order]

        self.mkdir(self.path + "/RESULT/SCM/WM")
        PurchaseInResult_sheet.to_excel(f'{self.path}/RESULT/SCM/WM/仓库出入库及时率.xlsx', sheet_name="仓库入库及时率", index=False)
        book = load_workbook(f'{self.path}/RESULT/SCM/WM/仓库出入库及时率.xlsx')
        writer = pd.ExcelWriter(f"{self.path}/RESULT/SCM/WM/仓库出入库及时率.xlsx", engine='openpyxl')
        writer.book = book
        self.PurchaseInData.to_excel(writer, "当月仓库入库情况", index=False)
        MaterialOutResult_sheet.to_excel(writer, "仓库出库及时率", index=False)
        self.MaterialOutData.to_excel(writer, "当月仓库出库情况", index=False)
        writer.save()

    # def MaterialOut(self, MaterialOutResult_sheet):  # 仓库出库及时率
    #
    #     book = load_workbook(f'{self.path}/RESULT/SCM/WM/仓库出入库及时率.xlsx')
    #     writer = pd.ExcelWriter(f"{self.path}/RESULT/SCM/WM/仓库出入库及时率.xlsx", engine='openpyxl')
    #     writer.book = book
    #     MaterialOutResult_sheet.to_excel(writer, "仓库出库及时率", index=False)
    #     self.MaterialOutData.to_excel(writer, "当月仓库出库情况", index=False)
    #     writer.save()

    def run(self):
        self.GetWarehouse()
        # self.PurchaseIn()
        # self.MaterialOut()


if __name__ == '__main__':
    W = Warehouse()
    W.run()

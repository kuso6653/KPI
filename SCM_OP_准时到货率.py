import pandas as pd
from numpy import datetime64
from openpyxl import load_workbook

import Func


class ArriveTime:
    def __init__(self):
        self.func = Func
        self.ThisMonthStart, self.ThisMonthEnd, self.LastMonthEnd, self.LastMonthStart = self.func.GetDate()
        self.path = Func.Path()
        # 将上月首尾日期切割
        self.ThisMonthStart = str(self.ThisMonthStart).split(" ")[0].replace("-", "")
        # self.ThisMonthEnd = str(self.ThisMonthEnd).split(" ")[0].replace("-", "")
        self.LastMonthStart = str(self.LastMonthStart).split(" ")[0].replace("-", "")
        self.LastMonthEnd = str(self.LastMonthEnd).split(" ")[0].replace("-", "")
        self.PurchaseInData = pd.read_excel(
            f"{self.path}/DATA/SCM/OP/采购订单列表.XLSX",
            usecols=['订单编号', '行号', '实际到货日期', '制单人', '行关闭人'],
            converters={'行号': int, '实际到货日期': datetime64})
        self.PurchaseInData = self.PurchaseInData.rename(columns={'制单人': '采购员'})
        self.PurchaseInData = self.PurchaseInData.loc[self.PurchaseInData["行关闭人"].isnull()]  # 筛出未关闭的行

        self.Prescription = pd.read_excel(
            f"{self.path}/DATA/SCM/采购时效性统计表.XLSX",
            usecols=[0, 1, 6, 7, 9, 11, 12, 15, 16, 17], header=2,
            names=["行号", "采购订单号", "存货编码", "存货名称", "计划到货日期", "采购订单制单时间", "采购订单审核时间",
                   "到货单号", "到货单行号", "到货单制单时间"],
            converters={'计划到货日期': datetime64, '采购订单制单时间': datetime64, '采购订单审核时间': datetime64, '到货单制单时间': datetime64})

        self.SubcontractData = pd.read_excel(f"{self.path}/DATA/SCM/OP/委外加工单列表.XLSX",
                                             usecols=['单据编号', '委外加工单行号', '委外商', '生产订单号', '生产订单行号', '生产批号', '计划到货日期',
                                                      '实际到货日期', '存货编码', '存货名称', '工序行号', '委外工序', '制单人', '制单时间', '审核时间',
                                                      '行关闭人'],
                                             converters={'单据编号': str, '生产订单行号': str, '存货编码': str, '工序行号': str,
                                                         '制单时间': datetime64, '审核时间': datetime64, '委外加工单行号': str,
                                                         '实际到货日期': datetime64})
        self.SubcontractData = self.SubcontractData.rename(columns={'制单人': '采购员', '单据编号': '委外加工单号', '制单时间': '委外加工单制单时间',
                                                                    '审核时间': '委外加工单审核时间'})
        self.SubcontractData = self.SubcontractData.loc[self.SubcontractData["行关闭人"].isnull()]  # 筛出未关闭的行

        self.SubcontractGR = pd.read_excel(f"{self.path}/DATA/SCM/OP/委外收料列表.XLSX",
                                           usecols=['单据编号', '委外加工单号', '委外加工单行号', '是否检验', '是否检验完成', '制单人', '制单时间',
                                                    '审核时间'],
                                           converters={'单据编号': str, '委外加工单号': str, '委外加工单行号': str, '制单时间': datetime64,
                                                       '审核时间': datetime64, '实际到货日期': datetime64})
        self.SubcontractGR = self.SubcontractGR.rename(columns={'制单人': '收料单制单人', '单据编号': '委外收料单号', '制单时间': '委外收料单制单时间',
                                                                '审核时间': '委外收料单审核时间'})

    def mkdir(self, path):
        self.func.mkdir(path)

    def GetThisMonthArriveTime(self):  # 当月准时到货率 和 当月未到货清单
        self.Prescription = self.Prescription.dropna(subset=['行号'])  # 去除nan的列
        self.PurchaseInData = self.PurchaseInData.dropna(subset=['行号'])  # 去除nan的列
        self.PurchaseInData = self.PurchaseInData.rename(columns={'订单编号': '采购订单号', '行号': '采购订单行号'})
        self.Prescription = self.Prescription.rename(columns={'行号': '采购订单行号'})
        ThisMonthArriveData = self.Prescription[self.ThisMonthEnd >= self.Prescription['计划到货日期']]
        ThisMonthArriveData = ThisMonthArriveData[ThisMonthArriveData['计划到货日期'] >= self.ThisMonthStart]
        ThisMonthArriveData = pd.merge(ThisMonthArriveData, self.PurchaseInData, on=['采购订单号', '采购订单行号'])

        # 筛选 实际到货日期 为空的， 用 计划到货日期 补全
        ThisMonthArriveData['实际到货日期'] = ThisMonthArriveData['实际到货日期'].fillna(ThisMonthArriveData['计划到货日期'])
        ThisMonthArriveData['实际到货日期'] = pd.to_datetime(ThisMonthArriveData['实际到货日期'].astype(str)) + pd.to_timedelta(
            '20:00:00')
        ThisMonthArriveData['计划到货日期'] = pd.to_datetime(ThisMonthArriveData['计划到货日期'].astype(str)) + pd.to_timedelta(
            '20:00:00')

        ThisMonthNoArriveData = ThisMonthArriveData[ThisMonthArriveData['到货单号'].isnull()]  # 筛选出未到货的行
        ThisMonthArriveData = ThisMonthArriveData[ThisMonthArriveData['到货单号'].notnull()]  # 筛选出已到货的行
        ThisMonthArriveData["实际审批延时/H"] = (
                (ThisMonthArriveData["到货单制单时间"] - ThisMonthArriveData["实际到货日期"]) / pd.Timedelta(1, 'H')).astype(int)
        ThisMonthArriveData.loc[ThisMonthArriveData["实际审批延时/H"] > 72, "实际单据状态"] = "逾期"
        ThisMonthArriveData.loc[ThisMonthArriveData["实际审批延时/H"] <= 72, "实际单据状态"] = "正常"
        ThisMonthArriveData.loc[ThisMonthArriveData["实际审批延时/H"] < 0, "实际单据状态"] = "提前"

        ThisMonthArriveData["计划审批延时/H"] = (
                (ThisMonthArriveData["到货单制单时间"] - ThisMonthArriveData["计划到货日期"]) / pd.Timedelta(1, 'H')).astype(int)

        ThisMonthArriveData.loc[ThisMonthArriveData["计划审批延时/H"] > 72, "计划单据状态"] = "逾期"
        ThisMonthArriveData.loc[ThisMonthArriveData["计划审批延时/H"] <= 72, "计划单据状态"] = "正常"
        ThisMonthArriveData.loc[ThisMonthArriveData["计划审批延时/H"] < 0, "计划单据状态"] = "提前"

        # 委外订单统计
        ThisMonthSubcontractData = self.SubcontractData[self.ThisMonthEnd >= self.SubcontractData['计划到货日期']]
        ThisMonthSubcontractData = ThisMonthSubcontractData[ThisMonthSubcontractData['计划到货日期'] >= self.ThisMonthStart]
        ThisMonthSubcontractData = pd.merge(ThisMonthSubcontractData, self.SubcontractGR, on=['委外加工单号', '委外加工单行号'])

        ThisMonthSubcontractData['实际到货日期'] = ThisMonthSubcontractData['实际到货日期'].fillna(
            ThisMonthSubcontractData['计划到货日期'])
        ThisMonthSubcontractData['实际到货日期'] = pd.to_datetime(
            ThisMonthSubcontractData['实际到货日期'].astype(str)) + pd.to_timedelta(
            '20:00:00')

        ThisMonthSubcontractData['计划到货日期'] = pd.to_datetime(
            ThisMonthSubcontractData['计划到货日期'].astype(str)) + pd.to_timedelta(
            '20:00:00')
        ThisMonthSubcontractData["实际审批延时/H"] = (
                (ThisMonthSubcontractData["委外收料单制单时间"] - ThisMonthSubcontractData["实际到货日期"]) / pd.Timedelta(1,
                                                                                                            'H')).astype(
            int)
        ThisMonthSubcontractData.loc[ThisMonthSubcontractData["实际审批延时/H"] > 72, "实际单据状态"] = "逾期"
        ThisMonthSubcontractData.loc[ThisMonthSubcontractData["实际审批延时/H"] <= 72, "实际单据状态"] = "正常"
        ThisMonthSubcontractData.loc[ThisMonthSubcontractData["实际审批延时/H"] < 0, "实际单据状态"] = "提前"

        ThisMonthSubcontractData["计划审批延时/H"] = (
                (ThisMonthSubcontractData["委外收料单制单时间"] - ThisMonthSubcontractData["计划到货日期"]) / pd.Timedelta(1,
                                                                                                            'H')).astype(
            int)

        ThisMonthSubcontractData.loc[ThisMonthSubcontractData["计划审批延时/H"] > 72, "计划单据状态"] = "逾期"
        ThisMonthSubcontractData.loc[ThisMonthSubcontractData["计划审批延时/H"] <= 72, "计划单据状态"] = "正常"
        ThisMonthSubcontractData.loc[ThisMonthSubcontractData["计划审批延时/H"] < 0, "计划单据状态"] = "提前"

        ThisMonthArriveData_Order = ['采购订单号', '采购订单行号', '存货编码', '存货名称', '采购员', '计划到货日期', '实际到货日期', '采购订单制单时间',
                                     '采购订单审核时间', '到货单号', '到货单行号', '到货单制单时间', '实际审批延时/H', '实际单据状态', '计划审批延时/H',
                                     '计划单据状态']
        ThisMonthNoArriveData_Order = ['采购订单号', '采购订单行号', '存货编码', '存货名称', '采购员', '计划到货日期', '实际到货日期', '采购订单制单时间',
                                       '采购订单审核时间', '到货单号', '到货单行号', '到货单制单时间']
        ThisMonthSubcontractData_Order = ['委外加工单号', '委外加工单行号', '委外商', '采购员', '生产订单号', '生产订单行号', '生产批号', '存货编码',
                                          '存货名称', '工序行号', '委外工序', '计划到货日期', '实际到货日期', '委外加工单制单时间', '委外加工单审核时间',
                                          '委外收料单号', '收料单制单人', '是否检验', '是否检验完成', '委外收料单制单时间', '委外收料单审核时间',
                                          '实际审批延时/H', '实际单据状态', '计划审批延时/H', '计划单据状态']
        ThisMonthArriveData = ThisMonthArriveData[ThisMonthArriveData_Order]
        ThisMonthNoArriveData = ThisMonthNoArriveData[ThisMonthNoArriveData_Order]
        ThisMonthSubcontractData = ThisMonthSubcontractData[ThisMonthSubcontractData_Order]

        ThisMonthNoArriveData = ThisMonthNoArriveData[
            ThisMonthNoArriveData['实际到货日期'] <= datetime64(self.ThisMonthEnd)]  # 筛选出本月的单据

        #  采购订单到货及时率计算
        try:
            ExceedTimeActData = ThisMonthArriveData["实际单据状态"].value_counts()['逾期']
        except:
            ExceedTimeActData = 0
        try:
            ExceedTimePlanData = ThisMonthArriveData["计划单据状态"].value_counts()['逾期']
        except:
            ExceedTimePlanData = 0
        try:
            Count = ThisMonthNoArriveData.shape[0]
        except:
            Count = 0
        CountAll = ThisMonthArriveData.shape[0] + ThisMonthNoArriveData.shape[0]
        ActResult = format(float(1 - (ExceedTimeActData + Count) / CountAll), '.2%')
        PlanResult = format(float(1 - (ExceedTimePlanData + Count) / CountAll), '.2%')
        dict = {'采购订单实际逾期单据数': [ExceedTimeActData], '采购订单计划逾期单据数': [ExceedTimePlanData], '采购订单当月未到单据数': [Count],
                '采购订单当月订单总数': [CountAll], '采购订单当月实际准时到货率': [ActResult], '采购订单当月计划准时到货率': [PlanResult]}
        Result_Sheet = pd.DataFrame(dict)

        #  委外订单到货及时率计算
        try:
            ActSubCount = ThisMonthSubcontractData['实际单据状态'].value_counts()['逾期']
        except:
            ActSubCount = 0
        try:
            PlanSubCount = ThisMonthSubcontractData['计划单据状态'].value_counts()['逾期']
        except:
            PlanSubCount = 0
        SubCountAll = ThisMonthSubcontractData.shape[0]
        ActSubResult = format(float(1 - ActSubCount / SubCountAll), '.2%')
        PlanSubResult = format(float(1 - PlanSubCount / SubCountAll), '.2%')
        dict1 = {'委外订单实际逾期单据数': [ActSubCount], '委外订单计划逾期单据数': [PlanSubCount], '委外订单当月订单总数': [SubCountAll],
                 '委外订单当月实际准时到货率': [ActSubResult], '委外订单当月计划准时到货率': [PlanSubResult]}
        SubResult_Sheet = pd.DataFrame(dict1)

        self.mkdir(self.path + "/RESULT/SCM/OP")
        Result_Sheet.to_excel(f'{self.path}/RESULT/SCM/OP/准时到货率.xlsx', sheet_name="当月采购准时到货率", index=False)
        book = load_workbook(f'{self.path}/RESULT/SCM/OP/准时到货率.xlsx')
        writer = pd.ExcelWriter(f"{self.path}/RESULT/SCM/OP/准时到货率.xlsx", engine='openpyxl')
        writer.book = book
        ThisMonthArriveData.to_excel(writer, "当月采购已到货情况", index=False)
        ThisMonthNoArriveData.to_excel(writer, "当月采购未到货清单", index=False)
        SubResult_Sheet.to_excel(writer, "当月委外准时到货率", index=False)
        ThisMonthSubcontractData.to_excel(writer, "当月委外订货情况", index=False)
        writer.save()

    def GetHistoryMonthArriveTime(self):  # 历史未到货清单
        HistoryMonthArriveData = self.Prescription[self.Prescription['计划到货日期'] < self.ThisMonthStart]
        HistoryMonthArriveData = pd.merge(self.PurchaseInData, HistoryMonthArriveData, on=['采购订单号', '采购订单行号'])
        HistoryMonthArriveData["实际到货日期"][HistoryMonthArriveData["实际到货日期"].isnull()] = HistoryMonthArriveData['计划到货日期']
        # 当 采购订单审核时间 或 到货单制单时间 为空值的时候取其数值
        HistoryMonthArriveData = HistoryMonthArriveData[
            (HistoryMonthArriveData["采购订单审核时间"].isnull()) | (HistoryMonthArriveData["到货单制单时间"].isnull())]
        order = ['采购订单号', '采购订单行号', '存货编码', '存货名称', '采购员', '计划到货日期', '实际到货日期', '采购订单制单时间', '采购订单审核时间', '到货单号', '到货单行号',
                 '到货单制单时间']
        HistoryMonthArriveData = HistoryMonthArriveData[order]

        book = load_workbook(f'{self.path}/RESULT/SCM/OP/准时到货率.xlsx')
        writer = pd.ExcelWriter(f"{self.path}/RESULT/SCM/OP/准时到货率.xlsx", engine='openpyxl')
        writer.book = book
        HistoryMonthArriveData.to_excel(writer, "历史未到货清单", index=False)
        writer.save()

    def run(self):
        self.GetThisMonthArriveTime()
        self.GetHistoryMonthArriveTime()


if __name__ == '__main__':
    AT = ArriveTime()
    AT.run()

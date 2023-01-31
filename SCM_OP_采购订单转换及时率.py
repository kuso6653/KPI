import pandas as pd
import calendar
import datetime
from datetime import timedelta
from numpy import datetime64
from openpyxl import load_workbook

import Func

pd.set_option('display.max_columns', None)


class OrderConversion:
    def __init__(self):
        self.func = Func
        self.ThisMonthStart, self.ThisMonthEnd, self.LastMonthEnd, self.LastMonthStart = self.func.GetDate()
        self.path = Func.Path()
        self.MRPScreenList = []  # 筛选合并的mrp数据
        self.MRPNewDataList = []  # 本月所有的mrp数据
        self.GroupMRPList = []  # 分组后的mrp数据

        self.ThisMonthStart = str(self.ThisMonthStart).split(" ")[0]
        self.LastMonthStart = str(self.LastMonthStart).split(" ")[0]
        self.ThisMonthEnd = str(self.ThisMonthEnd).split(" ")[0]
        self.this_month_check = self.ThisMonthStart

        OtherThisMonthStart = str(self.ThisMonthStart).split(" ")[0].replace("-", "")
        OtherLastMonthStart = str(self.LastMonthStart).split(" ")[0].replace("-", "")
        OtherThisMonthEnd = str(self.ThisMonthEnd).split(" ")[0].replace("-", "")
        self.PRProcessData = pd.read_excel(f"{self.path}/DATA/SCM/OP/请购执行进度表.XLSX",
                                           usecols=[1, 3, 6, 7, 8, 9, 10, 14, 15, 16, 20, 23], header=4,
                                           names=["请购单号", "请购单审核日期", "请购单行号", "存货编码", "存货名称", "规格型号", "数量", "采购订单号", "采购订单行号",
                                                  "采购订单下单日期", "计划到货日期", "采购订单制单人"],
                                           converters={'请购单号': str, '请购单行号': str, '存货编码': str, '数量': float,
                                                       '采购订单号': str, '采购订单行号': str, '计划到货日期': datetime64,
                                                       '采购订单下单日期': datetime64, '请购单审核日期': datetime64})
        self.PROnlyData = self.PRProcessData.loc[self.PRProcessData["采购订单号"].isnull()]
        self.PROnlyData = self.PROnlyData.dropna(subset=['请购单号'])  # 去除nan的列
        self.PRWithinPOData = self.PRProcessData.loc[self.PRProcessData["采购订单号"].notnull()]

        self.po_data = pd.read_excel(f"{self.path}/DATA/SCM/OP/采购订单列表.XLSX",
                                     usecols=['订单编号', '存货编码', '行号', '行关闭人'],
                                     converters={'订单编号': str, '行号': str, '存货编码': str}
                                     )
        self.po_data = self.po_data.rename(columns={'订单编号': '采购订单号', '行号': '采购订单行号'})  # 重命名

        self.pr_data = pd.read_excel(f"{self.path}/DATA/SCM/OP/请购单列表.XLSX",
                                     usecols=['单据号', '存货编码', '行号', '行关闭人'],
                                     converters={'单据号': str, '行号': str, '存货编码': str}
                                     )
        self.pr_data = self.pr_data.rename(columns={'单据号': '请购单号', '行号': '请购单行号'})  # 重命名

    def mkdir(self, path):
        self.func.mkdir(path)

    def ContrastData(self, BaseData, NewBase):
        BaseData = BaseData.dropna(subset=['物料编码'])  # 去除nan的列
        NewBase = NewBase.dropna(subset=['物料编码'])  # 去除nan的列
        out_data = pd.merge(BaseData, NewBase.drop(labels=['抓取时间', '是否客供料'], axis=1),
                            on=['物料编码', '物料名称', '需求跟踪号', '需求跟踪行号', '物料属性'])
        self.MRPScreenList.append(out_data)

    def GetOrderConversion(self):
        # 获取截取这个月份、年、上个月

        this_month = self.ThisMonthEnd.split("-")[1]
        year = self.ThisMonthEnd.split("-")[0]
        last_month = self.LastMonthStart.split("-")[1]

        last_work_days = self.func.WorkDays(year, last_month)  # 获取上个月工作日
        this_work_days = self.func.WorkDays(year, this_month)  # 获取这个月工作日
        work_days = []  # 设置到上月3天
        work_days.extend(last_work_days[-3:])
        work_days.extend(this_work_days)  # 将上个月最后三天和这个月工作日相合并
        work_days = self.func.ReformDays(work_days)  # 改造

        BaseDataList = []
        flag = 0
        for work_day in work_days:
            if flag < 3:
                try:
                    base_data = pd.read_excel(f"{self.path}/DATA/SCM/OP/MRP计划维护--全部{year}-{last_month}-{work_day}.XLSX",
                                              usecols=['物料编码', '物料名称', '需求跟踪号', '需求跟踪行号', '物料属性', '是否客供料'],
                                              converters={'物料编码': int, '需求跟踪行号': int}
                                              )
                    base_data = base_data.loc[base_data["物料属性"] == "采购"]
                    base_data = base_data.loc[base_data["是否客供料"].isnull()]
                    catch_data = datetime64(f"{year}-{last_month}-{work_day}")
                    base_data['抓取时间'] = catch_data
                    BaseDataList.append(base_data)
                    flag = flag + 1
                    continue
                except:
                    flag = flag + 1
                    continue
            else:
                try:
                    new_data = pd.read_excel(f"{self.path}/DATA/SCM/OP/MRP计划维护--全部{year}-{this_month}-{work_day}.XLSX",
                                             usecols=['物料编码', '物料名称', '需求跟踪号', '需求跟踪行号', '物料属性', '是否客供料'],
                                             converters={'物料编码': int, '需求跟踪行号': int}
                                             )
                    self.MRPNewDataList.append(new_data)
                except:
                    continue
                catch_data = datetime64(f"{year}-{this_month}-{work_day}")
                new_data['抓取时间'] = catch_data
                new_data = new_data.loc[new_data["物料属性"] == "采购"]
                new_data = new_data.loc[new_data["是否客供料"].isnull()]
                BaseDataList.append(new_data)  # 新添加新的base
                self.ContrastData(BaseDataList[0], new_data)  # 合并检查是否存在一样的
                del (BaseDataList[0])  # 删除第一个base
        all_mrp = pd.concat(self.MRPScreenList, axis=0, ignore_index=True)
        for name1, group in all_mrp.groupby(['物料编码', '需求跟踪号', '需求跟踪行号']):
            group = pd.DataFrame(group)  # 新建pandas
            group = group.sort_values(by='抓取时间', ascending=True)  # 升序排序
            max_data_list = group.head(1)  # 取最早的抓取时间，就是排序后的第一列
            self.GroupMRPList.append(max_data_list)  # 加入list
            flag = flag + 1

        self.GroupMRPData = pd.concat(self.GroupMRPList, axis=0, ignore_index=True)

    def ThisMonthNotConverted(self):  # 当月未转换MRP清单

        self.pr_data = self.pr_data.dropna(subset=['存货编码'])  # 去除nan的列
        self.po_data = self.po_data.dropna(subset=['存货编码'])  # 去除nan的列
        # pr-po 取差集
        PRNotConvertData = pd.merge(self.PROnlyData, self.pr_data, on=['请购单号', '存货编码', '请购单行号'], how="left")
        PRNotConvertData = PRNotConvertData.loc[PRNotConvertData['行关闭人'].isnull()]

        PRDelayPOData = pd.merge(self.PRWithinPOData, self.po_data, on=['采购订单号', '存货编码', '采购订单行号'], how="left")
        PRDelayPOData['下单延时/H'] = (
                (PRDelayPOData['采购订单下单日期'] - PRDelayPOData['请购单审核日期']) / pd.Timedelta(1, 'H')).astype(int)
        PRDelayPOData = PRDelayPOData.loc[PRDelayPOData['行关闭人'].isnull()]
        PRDelayPOData.loc[PRDelayPOData["下单延时/H"] > 48, "创建及时率"] = "超时"  # 计算出来的审批延时大于1天为超时
        PRDelayPOData.loc[PRDelayPOData["下单延时/H"] <= 48, "创建及时率"] = "正常"  # 小于等于1天为正常
        # self.pr_data = self.pr_data.append(self.po_data)
        # self.pr_data.drop_duplicates(keep=False, inplace=True)
        # self.pr_data.reset_index()

        #  当月大于2天的未转换MRP数据筛选
        MRPNow_data = self.GroupMRPData[self.GroupMRPData['抓取时间'] >= datetime64(self.this_month_check)]
        self.mkdir(self.path + '/RESULT/SCM/OP')
        PRNotConvertData.to_excel(f'{self.path}/RESULT/SCM/OP/采购订单转换及时率.xlsx', sheet_name="当月未转换PR清单", index=False)

        book = load_workbook(f'{self.path}/RESULT/SCM/OP/采购订单转换及时率.xlsx')
        writer = pd.ExcelWriter(f"{self.path}/RESULT/SCM/OP/采购订单转换及时率.xlsx", engine='openpyxl')
        writer.book = book
        PRDelayPOData.to_excel(writer, "历史转换PR清单", index=False)
        MRPNow_data.to_excel(writer, "当月未转换MRP清单", index=False)
        writer.save()

    def HistoryNotConverted(self):  # 历史未转换MRP清单
        self.mkdir(self.path + '/RESULT/SCM/OP')
        #  小于当月的历史未转换MRP数据筛选
        MRPHistory_data = self.GroupMRPData[self.GroupMRPData['抓取时间'] < datetime64(self.this_month_check)]
        book = load_workbook(f'{self.path}/RESULT/SCM/OP/采购订单转换及时率.xlsx')
        writer = pd.ExcelWriter(f"{self.path}/RESULT/SCM/OP/采购订单转换及时率.xlsx", engine='openpyxl')
        writer.book = book
        MRPHistory_data.to_excel(writer, "历史未转换MRP清单", index=False)
        writer.save()

    def run(self):
        self.GetOrderConversion()
        self.ThisMonthNotConverted()
        self.HistoryNotConverted()


if __name__ == '__main__':
    OC = OrderConversion()
    OC.run()

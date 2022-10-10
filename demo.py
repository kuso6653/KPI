import pandas as pd
import calendar
import datetime
from datetime import timedelta
from numpy import datetime64
from openpyxl import load_workbook

import Func

pd.set_option('display.max_columns', None)


class OrderConversion:
    def __init__(self):
        self.func = Func
        self.ThisMonthStart, self.ThisMonthEnd, self.LastMonthEnd, self.LastMonthStart = self.func.GetDate()
        self.path = Func.Path()
        self.MRPScreenList = []  # 筛选合并的mrp数据
        self.MRPNewDataList = []  # 本月所有的mrp数据
        self.GroupMRPList = []  # 分组后的mrp数据

        self.ThisMonthStart = str(self.ThisMonthStart).split(" ")[0]
        self.LastMonthStart = str(self.LastMonthStart).split(" ")[0]
        self.ThisMonthEnd = str(self.ThisMonthEnd).split(" ")[0]
        self.this_month_check = self.ThisMonthStart

        OtherThisMonthStart = str(self.ThisMonthStart).split(" ")[0].replace("-", "")
        OtherLastMonthStart = str(self.LastMonthStart).split(" ")[0].replace("-", "")
        OtherThisMonthEnd = str(self.ThisMonthEnd).split(" ")[0].replace("-", "")
        self.PRProcessData = pd.read_excel(f"F:/IT Documents/Office Documents/Project Documents/KPI考核/KPI数据分析/2022年1-8月份/基础数据/请购执行进度表.XLSX",
                                           usecols=[1, 3, 6, 7, 8, 9, 10, 14, 15, 16, 20, 23], header=4,
                                           names=["请购单号", "请购单审核日期", "请购单行号", "存货编码", "存货名称", "规格型号", "数量", "采购订单号",
                                                  "采购订单行号", "采购订单下单日期", "计划到货日期", "采购订单制单人"],
                                           converters={'请购单号': str, '请购单行号': str, '存货编码': str, '数量': float,
                                                       '采购订单号': str, '采购订单行号': str, '计划到货日期': datetime64,
                                                       '采购订单下单日期': datetime64, '请购单审核日期': datetime64})
        self.PROnlyData = self.PRProcessData.loc[self.PRProcessData["采购订单号"].isnull()]
        self.PROnlyData = self.PROnlyData.dropna(subset=['请购单号'])  # 去除nan的列
        self.PRWithinPOData = self.PRProcessData.loc[self.PRProcessData["采购订单号"].notnull()]

        self.po_data = pd.read_excel(f"F:/IT Documents/Office Documents/Project Documents/KPI考核/KPI数据分析/2022年1-8月份/基础数据/采购订单列表.XLSX",
                                     usecols=['订单编号', '存货编码', '行号', '行关闭人', '实际到货日期'],
                                     converters={'订单编号': str, '行号': int, '存货编码': int, '实际到货日期': datetime64}
                                     )
        self.po_data = self.po_data.rename(columns={'订单编号': '采购订单号', '行号': '采购订单行号'})  # 重命名

        self.pr_data = pd.read_excel(f"F:/IT Documents/Office Documents/Project Documents/KPI考核/KPI数据分析/2022年1-8月份/基础数据/请购单列表.XLSX",
                                     usecols=['单据号', '存货编码', '行号', '行关闭人'],
                                     converters={'单据号': str, '行号': int, '存货编码': int}
                                     )
        self.pr_data = self.pr_data.rename(columns={'单据号': '请购单号', '行号': '请购单行号'})  # 重命名

    def ThisMonthNotConverted(self):  # 当月未转换MRP清单

        self.pr_data = self.pr_data.dropna(subset=['存货编码'])  # 去除nan的列
        self.po_data = self.po_data.dropna(subset=['存货编码'])  # 去除nan的列
        # pr-po 取差集
        PRConvertData = pd.merge(self.PROnlyData, self.pr_data, on=['请购单号', '存货编码', '请购单行号'], how="left")
        PRNotConvertData = PRConvertData.loc[PRConvertData['行关闭人'].isnull()]
        PRCloseData = PRConvertData.loc[PRConvertData['行关闭人'].notnull()]

        PRDelayPOData = pd.merge(self.PRWithinPOData, self.po_data, on=['采购订单号', '存货编码', '采购订单行号'], how="left")
        PRDelayPOData['下单延时/H'] = (
                (PRDelayPOData['采购订单下单日期'] - PRDelayPOData['请购单审核日期']) / pd.Timedelta(1, 'H')).astype(int)
        PRDelayPOData = PRDelayPOData.loc[PRDelayPOData['行关闭人'].isnull()]
        PRDelayPOData.loc[PRDelayPOData["下单延时/H"] > 48, "创建及时率"] = "超时"  # 计算出来的审批延时大于1天为超时
        PRDelayPOData.loc[PRDelayPOData["下单延时/H"] <= 48, "创建及时率"] = "正常"  # 小于等于1天为正常
        # self.pr_data = self.pr_data.append(self.po_data)
        # self.pr_data.drop_duplicates(keep=False, inplace=True)
        # self.pr_data.reset_index()

        PRNotConvertData.to_excel(f'./采购订单转换及时率.xlsx', sheet_name="未转换PR清单", index=False)

        book = load_workbook(f'./采购订单转换及时率.xlsx')
        writer = pd.ExcelWriter(f"./采购订单转换及时率.xlsx", engine='openpyxl')
        writer.book = book
        PRCloseData.to_excel(writer, "关闭的PR清单", index=False)
        PRDelayPOData.to_excel(writer, "历史转换PR清单", index=False)
        writer.save()

    # def HistoryNotConverted(self):  # 历史未转换MRP清单
    #     # self.mkdir(self.path + '/RESULT/SCM/OP')
    #     #  小于当月的历史未转换MRP数据筛选
    #     MRPHistory_data = self.GroupMRPData[self.GroupMRPData['抓取时间'] < datetime64(self.this_month_check)]
    #     book = load_workbook(f'{self.path}/RESULT/SCM/OP/采购订单转换及时率.xlsx')
    #     writer = pd.ExcelWriter(f"{self.path}/RESULT/SCM/OP/采购订单转换及时率.xlsx", engine='openpyxl')
    #     writer.book = book
    #     MRPHistory_data.to_excel(writer, "历史未转换MRP清单", index=False)
    #     writer.save()

    def run(self):
        # self.GetOrderConversion()
        self.ThisMonthNotConverted()
        # self.HistoryNotConverted()


if __name__ == '__main__':
    OC = OrderConversion()
    OC.run()

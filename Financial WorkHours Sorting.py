from decimal import Decimal

import openpyxl
import pandas as pd
from numpy import datetime64
from openpyxl import load_workbook, Workbook
import Func

order = ['生产订单', '行号', '生产批号', '物料编码', '物料名称', '单据号码', '报工单表头备注', '制单人', '制单日期', '标准工艺版本号',
         '标准工艺说明', '工序行号', '标准工序', '标准工序名称', '工作中心', '工作中心名称', '报工工序', '报工工序名称', '班组编码', '班组名称',
         '员工代号', '员工姓名', '生产数量', '合格数量', '单位标准工时', '总标准工时', '资源工时1', '资源名称1', '资源工时2', '资源名称2',
         '备注工时', '备注设备']


class WorkHour:
    def __init__(self):
        self.wb = Workbook()

        self.func = Func
        self.ThisMonthStart, self.ThisMonthEnd, self.LastMonthEnd, self.LastMonthStart = self.func.GetDate()
        self.path = Func.Path()
        self.work_list = ['打磨', '大、小件划线', '焊接', '冷作', '修毛刺', '人工工时', '水压试验', '攻螺纹、修毛刺、清洁', '清洁']
        self.ThisMonthStart = str(self.ThisMonthStart).split(" ")[0].replace("-", "")
        self.ThisMonthEnd = str(self.ThisMonthEnd).split(" ")[0].replace("-", "")
        self.PlanHourData = pd.DataFrame
        self.ProductionData = pd.read_excel(f"./DATA/生产订单列表.XLSX",
                                            usecols=['生产订单号', '行号', '物料编码', '工艺路线版本号'],
                                            converters={'生产订单号': str, '物料编码': str, '工艺路线版本号': str, '行号': str
                                                        })

        self.Routing_data = pd.read_excel(f"./DATA/工艺路线资料表--含资源.xlsx",
                                          header=3, usecols=["物料编码", "工作中心", "工作中心名称", "版本日期", "版本代号", "资源名称",
                                                             "版本说明", "工时(分子)", "工序行号", "工序说明", "工序代号"],
                                          converters={'物料编码': str, "工时(分子)": int, "版本代号": str, "工序行号": str})

        self.WorkHourData = pd.read_excel(f"./DATA/报工列表.xlsx",
                                          usecols=["制单日期", "单据号码", "制单人", "生产批号", "物料名称", "表头备注", "生产订单", "行号",
                                                   "物料编码", "生产数量", "资源工时1", "资源名称1", "资源工时2", "资源名称2", "移入工序行号",
                                                   "移入标准工序", "移入工序说明", "合格数量", "备注工时", "备注设备",
                                                   "班组编码", "班组名称", "员工代号", "员工姓名"],
                                          converters={'单据号码': str, '行号': str, "资源工时1": int, "资源工时2": int, "备注工时": str,
                                                      "生产数量": int, '物料编码': str, '移入工序行号': str,
                                                      '员工代号': str, "表头备注": str, "生产订单": str, "物料名称": str, "生产批号": str})
        # d.fillna(value=0)

        self.WorkHourData = self.WorkHourData.rename(columns={'移入工序行号': '工序行号', '移入标准工序': '报工工序', '移入工序说明': '报工工序名称',
                                                              '表头备注': '报工单表头备注'})
        self.Routing_data = self.Routing_data.rename(columns={'工时(分子)': '单位标准工时', '版本代号': '标准工艺版本号', '版本说明': '标准工艺说明',
                                                              '工序说明': '标准工序名称', '工序代号': '标准工序'})
        self.ProductionData = self.ProductionData.rename(columns={'生产订单号': '生产订单', '工艺路线版本号': '标准工艺版本号'})

        self.ProductionData = self.ProductionData.dropna(subset=['标准工艺版本号'])  # 去除nan的列

        self.RoutingArtificial = self.Routing_data[self.Routing_data['资源名称'].isin(self.work_list)]  # 筛选工艺路线中人工工时
        self.RoutingNonArtificial = self.Routing_data[
            -self.Routing_data['资源名称'].isin(self.work_list)]  # 筛选工艺路线非人工(机械工时)

        # self.RoutingArtificial.to_excel(f'./RESULT/RoutingArtificial.xlsx' ,index=False)
        # self.RoutingNonArtificial.to_excel(f'./RESULT/RoutingNonArtificial.xlsx' ,index=False)
        # self.ProductionData.to_excel(f'./RESULT/ProductionData.xlsx' ,index=False)

        # 工艺路线 和 生产订单列表 进行合并
        self.RoutingArtificial = pd.merge(self.ProductionData, self.RoutingArtificial, on=['标准工艺版本号', '物料编码'],
                                          how="right")
        self.RoutingNonArtificial = pd.merge(self.ProductionData, self.RoutingNonArtificial, on=['标准工艺版本号', '物料编码'],
                                             how="right")
        # 上两个表合并后 与 报工列表 进行合并
        self.RoutingArtificial = pd.merge(self.RoutingArtificial, self.WorkHourData, how="right",
                                          on=['物料编码', '生产订单', '行号', '工序行号'])
        self.RoutingNonArtificial = pd.merge(self.RoutingNonArtificial, self.WorkHourData, how="right",
                                             on=['物料编码', '生产订单', '行号', '工序行号'])
        # 再将人工和非人工表进行合并
        self.AllData = pd.concat([self.RoutingArtificial, self.RoutingNonArtificial], axis=0, ignore_index=True)
        self.AllData["总标准工时"] = self.AllData["合格数量"] * self.AllData["单位标准工时"]  # 计算 标准工时
        self.AllData = self.AllData[order]

    def GetWorkData(self, group, name):
        # group = pd.merge(group, self.StandardData, on=['物料编码', '工序行号', '标准工序', '版本代号'])
        group["总标准工时"] = group["合格数量"] * group["单位标准工时"]  # 计算 标准工时
        del group["资源名称"]
        del group["版本代号"]
        del group["版本日期"]
        del group["资源工时1"]
        del group["资源工时2"]
        del group["资源名称1"]
        del group["资源名称2"]
        # group = group.rename(columns={'生产批号': '项目号'})
        group = group[order]
        group = group.sort_values(by=['生产订单', '行号', '物料编码', '工序行号', '标准工序'], ascending=True)  # 升序排列
        group['报工部门'] = name
        return group

    def GetWorkHour(self):
        pass

    def mkdir(self, path):
        self.func.mkdir(path)

    def SaveSheet(self, data, name):  # 新建excel页签并保存数据
        SaveBook = load_workbook(f'{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx')
        writer = pd.ExcelWriter(f"{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx", engine='openpyxl')
        writer.book = SaveBook
        data.to_excel(writer, f"{name}", index=False)
        writer.save()

    def FullSaveSheet(self, name):  # 新建空excel页签并保存数据
        data = pd.DataFrame()
        SaveBook = load_workbook(f'{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx')
        writer = pd.ExcelWriter(f"{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx", engine='openpyxl')
        writer.book = SaveBook
        data.to_excel(writer, f"{name}", index=False)
        writer.save()

    def SaveFile(self):
        pass
        # try:
        #     self.SaveSheet(AllData, "总数据")
        # except:
        #     self.FullSaveSheet("总数据")
        #
        # workbook = openpyxl.load_workbook(f'{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx')
        # del workbook["Sheet"]
        # workbook.save(f'{self.path}/RESULT/WORKHOUR/报工工时统计.xlsx')

    def run(self):
        self.AllData.to_excel(f'./RESULT/报工预警表.xlsx', index=False)


if __name__ == '__main__':
    WH = WorkHour()
    WH.run()
